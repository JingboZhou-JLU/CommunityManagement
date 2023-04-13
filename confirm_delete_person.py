
# 获取xlxs表格信息
import csv

from openpyxl import load_workbook

wb = load_workbook('data/9.15黎明居民档案-总表.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet
sheet1 = sheets[0]

all_valid_person = []
person = []
cnt = -1

for col in sheet1['F']:  # 居民档案姓名
    person.clear()
    cnt = cnt + 1
    if cnt == 0:
        continue
    person.append(col.value)
    all_valid_person.append(person.copy())

cnt = -1
for col in sheet1['G']:  # 居民身份证号
    cnt = cnt + 1
    if cnt == 0:
        continue
    all_valid_person[cnt - 1].append(col.value)

cnt = -1
for col in sheet1['H']:  # 居民性别
    cnt = cnt + 1
    if cnt == 0:
        continue
    all_valid_person[cnt - 1].append(col.value)

"""
获取网站上个人信息
"""

# 获取xlxs表格信息
wb = load_workbook('data/大庆高新区一人一档基本信息台账20220207121551.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet
sheet1 = sheets[0]

all_person = []
person = []
cnt = -1
for col in sheet1['F']:  # 一人一档姓名
    person.clear()
    cnt = cnt + 1
    if cnt == 0:
        continue
    person.append(col.value)
    all_person.append(person.copy())

cnt = -1
for col in sheet1['G']:  # 一人一档身份证号
    cnt = cnt + 1
    if cnt == 0:
        continue
    all_person[cnt - 1].append(col.value)

cnt = -1
for col in sheet1['J']:  # 一人一档性别
    cnt = cnt + 1
    if cnt == 0:
        continue
    all_person[cnt - 1].append(col.value)

# 删除第一项
del all_person[0]

delete_person = []
add_person_name = []

name_list = []
for i in all_person:
    name_list.append(i[0])

name_list_valid = []
for i in all_valid_person:
    name_list_valid.append(i[0])

for per in all_person:
    if per[0] not in name_list_valid:
        delete_person.append(per.copy())

for name in name_list_valid:
    if name not in name_list:
        add_person_name.append(name)

add_person = []  # 完整的基本信息

# 获取xlxs表格信息
wb = load_workbook('data/9.15黎明居民档案-总表.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet
sheet1 = sheets[0]
rows = sheet1.rows
person_info = []


# 迭代读取所有的行
for row in rows:
    cnt = -1
    row_val = []
    for col in row:
        cnt = cnt + 1
        if cnt == 0:
            continue
        if cnt == 18:
            break
        row_val.append(col.value)
    if row_val[2] == '郭焕山':
        break
    print(row_val)
    person_info.append(row_val)


# 表项数目
print("一人一档数目", len(all_person))
print("在档数目", len(all_valid_person))
print("冗余人数", len(delete_person))
print("添加人数",len(add_person_name))

del person_info[0]
for per in person_info:
    if per[2] is None:
        index = person_info.index(per)       # 获取索引信息
        person_info[index][2] = person_info[index - 1][2]    # 修改户主信息

for per in person_info:
    if per[4] in add_person_name:   # 获取信息
        add_person.append(per)

# 写入csv文件
f = open('delete_person.csv', 'w', encoding='utf-8',newline="")
csv_writer = csv.writer(f)
csv_writer.writerow(["姓名", "身份证号", "性别"])
for per in delete_person:
    csv_writer.writerow(per)
f.close()


# f = open('add_person.csv', 'w', encoding='utf-8', newline="")
# csv_writer = csv.writer(f)
# for per in add_person:
#     csv_writer.writerow(per)
# f.close()
