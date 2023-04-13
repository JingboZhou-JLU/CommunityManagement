# 获取xlxs表格信息
import csv
from openpyxl import load_workbook

wb = load_workbook('data/9.15黎明居民档案-总表.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet
sheet1 = sheets[0]

person = []
cnt = -1

for col in sheet1['F']:  # 居民档案姓名
    cnt = cnt + 1
    if cnt == 0:
        continue
    person.append(col.value)


# print(person)
# print(len(person))

person_state = []
wb2 = load_workbook('data/黎明社区抗原发放行动居民台账5.7.xlsx')
sheets2 = wb2.worksheets  # 获取当前所有的sheet
sheet2 = sheets2[0]

cnt  = -1
for col in sheet2['B']:  # 居民档案姓名
    cnt = cnt + 1
    if cnt == 0 or cnt == 1:
        continue
    person_state.append([col.value,0])
    if col.value == '张月':
        break

cnt = -1
for col in sheet2['H']:  # 居民档案姓名
    cnt = cnt + 1
    if cnt == 0 or cnt == 1:
        continue
    if col.value == "已领":
        col.value = "常住"
    person_state[cnt-2][1] = col.value
    if cnt-2 == len(person_state) -1:
        break

# print(person_state)

state = []
flag  = 0
for per in person:
    flag = 0
    for i in person_state:
        if per == i[0]:
            state.append(i[1])
            flag = 1
            break
    if flag == 1:
        continue
    else:
        state.append('*')


# print(state)
# print(len(person),len(state))

cnt = 2
for i in state:
    str1 = 'V' + str(cnt)
    sheet1[str1].value = i
    cnt = cnt + 1


wb.save('data/new.xlsx')